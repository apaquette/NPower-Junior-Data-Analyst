"""_summary_
"""
import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
value_per_supplier = {}
under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_number = product_list.cell(product_row, 1).value

    inventory_price = product_list.cell(product_row, 5)

    # calculation of products per supplier
    if supplier in products_per_supplier:
        products_per_supplier[supplier] += 1
    else:
        products_per_supplier[supplier] = 1

    # total value of inventory per supplied
    if supplier in value_per_supplier:
        value_per_supplier[supplier] += inventory * price
    else:
        value_per_supplier[supplier] = inventory * price

    # logic products with inventory below 10
    if inventory < 10:
        under_10_inv[product_number] = inventory

    # add value for total inventory price
    inventory_price.value = inventory * price

# fix trailing values due to floating point precision
for supplier in value_per_supplier:
    value_per_supplier[supplier] = round(value_per_supplier.get(supplier), 2)

inv_file.save("inventory_with_total_value.xlsx")

print(products_per_supplier)
print(value_per_supplier)
print(under_10_inv)
